# -*- coding: utf-8 -*-
"""
Created on Sat Jun 19 17:28:06 2021

@author: admOS
"""

from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl
import datetime

def read_page(login, number):
    url = f"https://habr.com/ru/users/{login}/favorites/posts/page{number}/"
    page = urlopen(url)
    page_object = BeautifulSoup(page.read())
    return page_object

def convert_str_to_datetime(datetime_str):
    """
    Конвертирует строку с датой в формате 2013-09-10, 20:56 в объект datetime.
    """
    return datetime.datetime.strptime(datetime_str, "%Y-%m-%d, %H:%M")

def convert_datetime_to_str(datetime_obj):
    """
    Конвертирует объект datetime в строку с датой в формате 10 September 2013.
    """
    return datetime.datetime.strftime(datetime_obj, "%d %B %Y")

username = input("Введите ваше имя пользователя на habr.com: ")
xl = openpyxl.Workbook()
ws = xl.active
ws.title = "habr.com"
row = 1
page_obj = read_page(username, 1)
num_pages = int(page_obj.find("div", {"class": "tm-pagination__pages"}).findAll("a", {"class": "tm-pagination__page"})[-1].get_text().strip())
for num in range(1, num_pages+1):
    page_obj = read_page(username, num)
    articles = page_obj.findAll("article", {"class": "tm-articles-list__item"})
    for article in articles:
        date = article.find("time").attrs["title"]                             # время публикации поста
        date = convert_datetime_to_str(convert_str_to_datetime(date))
        title_obj = article.find("h2", {"class": "tm-title tm-title_h2"})
        # if title_obj:
        title = title_obj.get_text().strip()  # название поста
        link = "https://habr.com" + title_obj.find("a").attrs["href"]  # ссылка на пост
        try:
            hubs = article.find("div", {"class": "tm-publication-hubs"}).stripped_strings
            hubs = [hub for hub in hubs if all(["*" not in hub, "Блог компании" not in hub])]
        except AttributeError:
            hubs = []
        try:
            labels_obj = article.find("div", {"class": "tm-article-labels__container"}).stripped_strings
            labels = [label for label in labels_obj if label]
        except (AttributeError, TypeError):
            labels = []
        # else:
        #     title_obj = article.find("h2", {"class": "tm-megapost-snippet__title"})
        #     title = title_obj.get_text().strip()  # название поста
        #     link = "https://habr.com" + article.find("a", {"class": "tm-megapost-snippet__link tm-megapost-snippet__card"}).attrs["href"]  # ссылка на пост
        #     hubs_obj = article.find("div", {"class": "tm-publication-hubs"})
        #     labels_obj = article.find("div", {"class": "tm-article-labels__container"})
        ws.cell(row=row, column=1).style = "Hyperlink"
        ws.cell(row=row, column=1).value = title
        ws.cell(row=row, column=1).hyperlink = link
        ws.cell(row=row, column=2, value=", ".join(hubs))
        ws.cell(row=row, column=3, value=", ".join(labels))
        ws.cell(row=row, column=4, value=date)
        row +=1
ws.column_dimensions["A"].width = 130
ws.column_dimensions["B"].width = 112
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 20
xl.save("habr.com.xlsx")
xl.close()